from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from .models import CellType, CultureVessel
import json
import decimal
from decimal import Decimal
from .utils import validate_required_fields, perform_calculation
from django.conf import settings
from io import BytesIO
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
except Exception:  # Python < 3.9 fallback
    ZoneInfo = None

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import (
    Border as XLBorder,
    Side as XLSide,
    Alignment as XLAlignment,
    Font as XLFont,
)
from openpyxl.drawing.image import Image as XLImage

# PDF generation
from django.template.loader import render_to_string
import base64
import os
from django.contrib.staticfiles import finders


@login_required
def calculator_view(request):
    """Main calculator view"""
    from bitbio.countries import COUNTRIES

    return render(request, "calculator.html", {"countries": COUNTRIES})


@method_decorator(csrf_exempt, name="dispatch")
class CellTypesAPIView(View):
    """API endpoint for cell types"""

    def get(self, request):
        try:
            cell_types = list(
                CellType.objects.values("id", "product_name", "sku", "seeding_density")
            )
            return JsonResponse(cell_types, safe=False)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@method_decorator(csrf_exempt, name="dispatch")
class CultureVesselsAPIView(View):
    """API endpoint for culture vessels"""

    def get(self, request):
        try:
            vessels = list(
                CultureVessel.objects.values(
                    "id", "plate_format", "surface_area_cm2", "media_volume_per_well_ml"
                )
            )
            return JsonResponse(vessels, safe=False)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@method_decorator(csrf_exempt, name="dispatch")
class DownloadExcelView(View):
    """Handle Excel download"""

    def post(self, request):
        try:
            data = request.POST or {}

            # Helpers
            def strip_tags(value: str) -> str:
                from html import unescape
                import re

                if value is None:
                    return ""
                # remove HTML tags
                return unescape(re.sub(r"<[^>]*>", "", str(value)))

            def format_cell_count(value: str) -> str:
                if not value:
                    return ""
                raw = str(value).replace(",", "")
                import re

                m = re.match(r"([0-9.]+)[eE]([+-]?[0-9]+)", raw)
                if m:
                    base = float(m.group(1))
                    exp = int(m.group(2))
                    num = base * (10**exp)
                    return f"{num:,.0f}"
                m2 = re.match(r"([0-9.]+)\s*x\s*10\^?([0-9]+)", raw)
                if m2:
                    base = float(m2.group(1))
                    exp = int(m2.group(2))
                    num = base * (10**exp)
                    return f"{num:,.0f}"
                try:
                    return f"{float(raw):,}"
                except Exception:
                    return str(value)

            def unicode_superscript_to_int(s: str) -> int:
                mapping = {
                    "⁰": "0",
                    "¹": "1",
                    "²": "2",
                    "³": "3",
                    "⁴": "4",
                    "⁵": "5",
                    "⁶": "6",
                    "⁷": "7",
                    "⁸": "8",
                    "⁹": "9",
                }
                out = "".join(mapping.get(ch, "") for ch in s)
                return int(out) if out else 0

            def clean_scientific_notation(value: str) -> str:
                if not value:
                    return ""
                import re

                v = str(value)
                m_html = re.search(r"([0-9.]+)\s*x\s*10<sup>([+-]?[0-9]+)</sup>", v)
                if m_html:
                    base = float(m_html.group(1))
                    exp = int(m_html.group(2))
                    return f"{base * (10 ** exp):,.0f}"
                m_uni = re.search(r"([0-9.]+)\s*x\s*10([⁰¹²³⁴⁵⁶⁷⁸⁹]+)", v)
                if m_uni:
                    base = float(m_uni.group(1))
                    exp = unicode_superscript_to_int(m_uni.group(2))
                    return f"{base * (10 ** exp):,.0f}"
                m_e = re.search(r"([0-9.]+)[eE]([+-]?[0-9]+)", v)
                if m_e:
                    base = float(m_e.group(1))
                    exp = int(m_e.group(2))
                    return f"{base * (10 ** exp):,.0f}"
                try:
                    return f"{float(strip_tags(v)):,.0f}"
                except Exception:
                    return strip_tags(v)

            # Timezone
            tz_name = data.get("timezone") or getattr(settings, "TIME_ZONE", "UTC")
            tz = None
            if ZoneInfo is not None:
                try:
                    tz = ZoneInfo(tz_name)
                except Exception:
                    tz = ZoneInfo(getattr(settings, "TIME_ZONE", "UTC"))
            now = datetime.now(tz) if tz else datetime.now()
            date_part = now.strftime("%Y-%m-%d")
            time_part = now.strftime("%H-%M")
            formatted_timestamp = f"{date_part} - {time_part}"
            formatted_date = now.strftime("%Y-%m-%d %H:%M:%S")

            # Build workbook
            wb = Workbook()
            ws = wb.active

            # Default font
            # openpyxl has no global default setter; set per usage

            # Logo
            logo_ok = False
            # Try using Django's staticfiles finder first
            logo_found = finders.find("images/bitbio-logo.png")
            if logo_found and os.path.exists(logo_found):
                try:
                    img = XLImage(logo_found)
                    img.width = 120
                    img.height = 36
                    ws.add_image(img, "A1")  # add left margin by starting in column B
                    logo_ok = True
                except Exception:
                    logo_ok = False
            if not logo_ok:
                possible_paths = [
                    os.path.join(
                        settings.BASE_DIR,
                        "bitbio",
                        "static",
                        "images",
                        "bitbio-logo.png",
                    ),
                    os.path.join(
                        settings.BASE_DIR, "staticfiles", "images", "bitbio-logo.png"
                    ),
                ]
                for p in possible_paths:
                    if os.path.exists(p):
                        try:
                            img = XLImage(p)
                            img.width = 120
                            img.height = 36
                            ws.add_image(
                                img, "B1"
                            )  # add left margin by starting in column B
                            logo_ok = True
                            break
                        except Exception:
                            pass
            # If no logo is found, leave A1 empty (no fallback text)

            ws.row_dimensions[1].height = 50  # a bit more top padding for the logo

            # Title
            ws["A2"] = "Cell seeding calculator"
            ws.merge_cells("A2:C2")
            ws["A2"].font = XLFont(bold=True, size=14)
            ws["A2"].alignment = XLAlignment(horizontal="center")

            # Timestamp
            ws["A4"] = f"Generated on: {formatted_date}"
            ws.merge_cells("A4:C4")
            ws["A4"].font = XLFont(size=9)

            # Headers
            ws["A8"] = "Input data"
            ws["B8"] = "Value"
            ws["C8"] = "Unit"
            for c in ("A8", "B8", "C8"):
                ws[c].font = XLFont(bold=True, size=14)

            # Input rows
            ws["A9"] = "Cell stock volume"
            ws["B9"] = data.get("suspensionVolume", "")
            ws["C9"] = "mL"

            ws["A10"] = "Live cell count - 1"
            ws["B10"] = format_cell_count(data.get("count1", ""))
            ws["C10"] = "cells/mL"
            ws["A11"] = "Live cell count - 2"
            ws["B11"] = format_cell_count(data.get("count2", ""))
            ws["C11"] = "cells/mL"
            ws["A12"] = "Live cell count - 3"
            ws["B12"] = format_cell_count(data.get("count3", ""))
            ws["C12"] = "cells/mL"

            ws["A13"] = "Cell viability - 1"
            ws["B13"] = data.get("viability1", "")
            ws["C13"] = "%"
            ws["A14"] = "Cell viability - 2"
            ws["B14"] = data.get("viability2", "")
            ws["C14"] = "%"
            ws["A15"] = "Cell viability - 3"
            ws["B15"] = data.get("viability3", "")
            ws["C15"] = "%"

            ws["A16"] = "Cell type"
            ws["B16"] = data.get("cellType", "")
            ws["C16"] = ""
            ws["A17"] = "Seeding density"
            ws["B17"] = data.get("seedingDensity", "")
            ws["C17"] = "cells/cm²"
            ws["A18"] = "Culture vessel"
            ws["B18"] = data.get("cultureVessel", "")
            ws["C18"] = ""
            ws["A19"] = "Surface area"
            ws["B19"] = data.get("surfaceArea", "")
            ws["C19"] = "cm²/well"
            ws["A20"] = "Volume"
            ws["B20"] = data.get("mediaVolume", "")
            ws["C20"] = "mL/well"
            ws["A21"] = "Number of wells to seed"
            ws["B21"] = data.get("wellCount", "")
            ws["C21"] = "wells"
            ws["A22"] = "Dead volume allowance"
            ws["B22"] = data.get("buffer", "")
            ws["C22"] = "%"

            # Results header
            ws["A24"] = "Results"
            ws["B24"] = "Value"
            ws["C24"] = "Unit"
            for c in ("A24", "B24", "C24"):
                ws[c].font = XLFont(bold=True, size=14)

            ws["A25"] = "Volume of media for final seeding solution"
            ws["B25"] = data.get("volumeToDilute", "")
            ws["C25"] = "mL"
            ws["A26"] = "Cell stock volume for final seeding solution"
            ws["B26"] = data.get("volumeToSeed", "")
            ws["C26"] = "mL"
            ws["A27"] = "Cell density"
            ws["B27"] = clean_scientific_notation(data.get("cellDensity", ""))
            ws["C27"] = "cells/mL"
            ws["A28"] = "Required number of cells (total)"
            ws["B28"] = clean_scientific_notation(data.get("requiredCells", ""))
            ws["C28"] = "cells"
            ws["A29"] = "Required number of cells (per well)"
            ws["B29"] = strip_tags(data.get("cellsPerWell", ""))
            ws["C29"] = "cells"

            # Column widths
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 12

            # Borders
            thin = XLSide(style="thin", color="000000")

            def apply_border(range_ref: str):
                for row in ws[range_ref]:
                    for cell in row:
                        cell.border = XLBorder(
                            top=thin, left=thin, right=thin, bottom=thin
                        )

            apply_border("A9:C22")
            apply_border("A25:C29")

            # Alignments
            def align_right(range_ref: str):
                for row in ws[range_ref]:
                    for cell in row:
                        cell.alignment = XLAlignment(horizontal="right")

            align_right("B9:B22")
            align_right("B25:B29")

            # Font sizes for data cells
            def set_font_size(range_ref: str, size: int):
                for row in ws[range_ref]:
                    for cell in row:
                        cell.font = XLFont(size=size)

            set_font_size("A9:C22", 11)
            set_font_size("A25:C29", 11)

            # Footer
            ws["A33"] = (
                "By using this tool, you agree to bit.bio's Cell seeding calculator - Terms and Conditions."
            )
            ws["A34"] = f"bit.bio © {now.strftime('%Y')}"
            ws["A33"].font = XLFont(size=9)
            ws["A34"].font = XLFont(size=9)

            # Build response
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            filename = (
                f"bit.bio - Cell Seeding Calculation - {formatted_timestamp}.xlsx"
            )
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


# Updated DownloadPDFView with production fixes


@method_decorator(csrf_exempt, name="dispatch")
class DownloadPDFView(View):
    """Handle PDF download with production-ready fallbacks"""

    def get(self, request):
        """Handle GET requests by redirecting to calculator page"""
        from django.shortcuts import redirect

        return redirect("calculator:calculator")

    def post(self, request):
        try:
            import logging

            logger = logging.getLogger(__name__)
            logger.info("PDF download request started")

            data = request.POST or {}

            # Timezone and timestamp setup (existing code)
            tz_name = data.get("timezone") or getattr(settings, "TIME_ZONE", "UTC")
            tz = None
            if ZoneInfo is not None:
                try:
                    tz = ZoneInfo(tz_name)
                except Exception:
                    tz = ZoneInfo(getattr(settings, "TIME_ZONE", "UTC"))
            now = datetime.now(tz) if tz else datetime.now()
            formatted_date = now.strftime("%Y-%m-%d %H:%M:%S")
            date_part = now.strftime("%Y-%m-%d")
            time_part = now.strftime("%H-%M")
            formatted_timestamp = f"{date_part} - {time_part}"
            filename = f"bit.bio - Cell Seeding Calculation - {formatted_timestamp}.pdf"

            # Helper functions (existing code)
            def strip_tags(value: str) -> str:
                from html import unescape
                import re

                if value is None:
                    return ""
                return unescape(re.sub(r"<[^>]*>", "", str(value)))

            def unicode_sup_to_text(s: str) -> str:
                mapping = {
                    "⁰": "0",
                    "¹": "1",
                    "²": "2",
                    "³": "3",
                    "⁴": "4",
                    "⁵": "5",
                    "⁶": "6",
                    "⁷": "7",
                    "⁸": "8",
                    "⁹": "9",
                    "⁻": "-",
                    "⁺": "+",
                }
                return "".join(mapping.get(ch, "") for ch in s)

            def sci_to_html_sup(value: str) -> str:
                if not value:
                    return ""
                import re

                v = str(value)
                if re.search(r"([0-9.]+)\s*x\s*10<sup>([+-]?[0-9]+)</sup>", v):
                    return v
                m_uni = re.search(r"([0-9.]+)\s*x\s*10([⁰¹²³⁴⁵⁶⁷⁸⁹⁻⁺]+)", v)
                if m_uni:
                    base = m_uni.group(1)
                    exp = unicode_sup_to_text(m_uni.group(2))
                    return f"{base} x 10<sup>{exp}</sup>"
                m_e = re.search(r"([0-9.]+)[eE]([+-]?[0-9]+)", v)
                if m_e:
                    base = m_e.group(1)
                    exp = m_e.group(2).lstrip("+")
                    return f"{base} x 10<sup>{exp}</sup>"
                m_caret = re.search(r"([0-9.]+)\s*x\s*10\^([+-]?[0-9]+)", v)
                if m_caret:
                    base = m_caret.group(1)
                    exp = m_caret.group(2)
                    return f"{base} x 10<sup>{exp}</sup>"
                try:
                    return f"{float(strip_tags(v)):,}"
                except Exception:
                    return strip_tags(v)

            # Logo handling (existing code)
            logo_base64 = None
            logo_found = finders.find("images/bitbio-logo.png")
            if logo_found and os.path.exists(logo_found):
                try:
                    with open(logo_found, "rb") as f:
                        logo_base64 = "data:image/png;base64," + base64.b64encode(
                            f.read()
                        ).decode("ascii")
                except Exception:
                    logo_base64 = None

            if not logo_base64:
                logo_path_candidates = [
                    os.path.join(
                        settings.BASE_DIR,
                        "bitbio",
                        "static",
                        "images",
                        "bitbio-logo.png",
                    ),
                    os.path.join(
                        settings.BASE_DIR, "staticfiles", "images", "bitbio-logo.png"
                    ),
                ]
                for p in logo_path_candidates:
                    if os.path.exists(p):
                        try:
                            with open(p, "rb") as f:
                                logo_base64 = (
                                    "data:image/png;base64,"
                                    + base64.b64encode(f.read()).decode("ascii")
                                )
                                break
                        except Exception:
                            pass

            # Context preparation (existing code)
            context = {
                "current_year": now.strftime("%Y"),
                "timestamp": formatted_date,
                "logo_base64": logo_base64,
                "suspension_volume": data.get("suspensionVolume", ""),
                "count1": data.get("count1", ""),
                "count2": data.get("count2", ""),
                "count3": data.get("count3", ""),
                "viability1": data.get("viability1", ""),
                "viability2": data.get("viability2", ""),
                "viability3": data.get("viability3", ""),
                "cell_type": data.get("cellType", ""),
                "seeding_density": data.get("seedingDensity", ""),
                "culture_vessel": data.get("cultureVessel", ""),
                "surface_area": data.get("surfaceArea", ""),
                "media_volume": data.get("mediaVolume", ""),
                "well_count": data.get("wellCount", ""),
                "buffer": data.get("buffer", ""),
                "volume_to_dilute": data.get("volumeToDilute", ""),
                "volume_to_seed": data.get("volumeToSeed", ""),
                "cell_density_html": sci_to_html_sup(data.get("cellDensity", "")),
                "required_cells": strip_tags(data.get("requiredCells", "")),
                "required_cells_html": sci_to_html_sup(data.get("requiredCells", "")),
                "cells_per_well": strip_tags(data.get("cellsPerWell", "")),
                "volume_per_well": data.get("volumePerWell", ""),
                "warnings": data.get("warnings", ""),
            }

            # Render HTML template
            try:
                html = render_to_string("pdf_download.html", context)
            except Exception as template_error:
                logger.error(f"Template rendering failed: {str(template_error)}")
                return JsonResponse(
                    {"error": f"Template rendering failed: {str(template_error)}"},
                    status=500,
                )

            # Generate PDF using Playwright
            logger.info("Attempting PDF generation with Playwright")
            try:
                # Check if Playwright is available
                try:
                    from playwright.sync_api import sync_playwright

                    logger.info("Playwright imported successfully")
                except ImportError as e:
                    logger.error(f"Playwright not available: {str(e)}")
                    raise Exception(
                        "Playwright not available. Please install with: pip install playwright && playwright install chromium"
                    )

                # Enhanced Playwright configuration for production
                with sync_playwright() as p:
                    browser = None
                    try:
                        # Browser launch arguments for better stability
                        browser_args = [
                            "--no-sandbox",
                            "--disable-dev-shm-usage",
                            "--disable-gpu",
                            "--disable-web-security",
                            "--disable-features=VizDisplayCompositor",
                            "--disable-background-timer-throttling",
                            "--disable-backgrounding-occluded-windows",
                            "--disable-renderer-backgrounding",
                            "--disable-extensions",
                            "--disable-plugins",
                            "--disable-images",  # Faster rendering
                        ]

                        # Try chromium first (most reliable)
                        try:
                            browser = p.chromium.launch(
                                headless=True,
                                args=browser_args,
                                timeout=30000,
                            )
                            logger.info("Chromium browser launched successfully")
                        except Exception as chromium_error:
                            logger.warning(f"Chromium launch failed: {chromium_error}")
                            # Try firefox as fallback
                            try:
                                browser = p.firefox.launch(
                                    headless=True,
                                    timeout=30000,
                                )
                                logger.info("Firefox browser launched successfully")
                            except Exception as firefox_error:
                                logger.warning(
                                    f"Firefox launch failed: {firefox_error}"
                                )
                                raise Exception(
                                    "No browser available. Please install browser dependencies with: playwright install-deps"
                                )

                        if not browser:
                            raise Exception("Failed to launch any browser")

                        page = browser.new_page()

                        # Set viewport for consistent rendering
                        page.set_viewport_size({"width": 1200, "height": 800})

                        # Load content with timeout
                        page.set_content(html, wait_until="networkidle", timeout=30000)

                        # Wait for any dynamic content to load
                        page.wait_for_timeout(1000)  # 1 second buffer

                        # Generate PDF with optimized settings
                        pdf_bytes = page.pdf(
                            format="A4",
                            landscape=True,
                            margin={
                                "top": "10mm",
                                "right": "10mm",
                                "bottom": "10mm",
                                "left": "10mm",
                            },
                            print_background=True,
                            prefer_css_page_size=True,
                            scale=0.8,  # Slightly smaller scale for better fit
                        )

                        logger.info("PDF generated successfully with Playwright")

                        response = HttpResponse(
                            pdf_bytes, content_type="application/pdf"
                        )
                        response["Content-Disposition"] = (
                            f'attachment; filename="{filename}"'
                        )
                        return response

                    except Exception as browser_error:
                        if browser:
                            try:
                                browser.close()
                            except:
                                pass
                        raise browser_error

            except Exception as e:
                logger.error(f"Playwright PDF generation failed: {str(e)}")

                # Fallback: Return HTML with print instructions
                logger.warning("PDF generation failed, falling back to HTML")
                html_filename = filename.replace(".pdf", ".html")

                html_with_instructions = html.replace(
                    "<body>",
                    """<body>
                    <div style="background: #fff3cd; border: 1px solid #ffc107; padding: 15px; margin: 10px 0; border-radius: 5px; text-align: center;">
                        <h3 style="margin-top: 0; color: #856404;">PDF Generation Temporarily Unavailable</h3>
                        <p style="margin-bottom: 0;"><strong>To save as PDF:</strong> Use your browser's print function (Ctrl+P or Cmd+P) and select "Save as PDF" as the destination.</p>
                        <p style="margin-top: 10px; font-size: 12px; color: #666;">Error: """
                    + str(e)
                    + """</p>
                    </div>""",
                )

                response = HttpResponse(
                    html_with_instructions, content_type="text/html; charset=utf-8"
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="{html_filename}"'
                )
                logger.info(f"Returning HTML fallback: {html_filename}")
            return response

        except Exception as e:
            logger.error(f"Complete PDF generation failure: {str(e)}")
            return JsonResponse(
                {
                    "error": "PDF generation is temporarily unavailable. Please try again later or contact support.",
                    "details": str(e) if settings.DEBUG else None,
                },
                status=500,
            )


@method_decorator(csrf_exempt, name="dispatch")
class CalculateAPIView(View):
    """Server-side calculation endpoint to mirror client-side calculator.js logic"""

    def post(self, request):
        try:
            payload = json.loads(request.body or b"{}")

            validation = validate_required_fields(payload)
            if validation.has_errors:
                return JsonResponse(
                    {
                        "success": False,
                        "validation": {
                            "missingFields": validation.missing_fields,
                            "negativeValueFields": validation.negative_value_fields,
                            "percentageOverLimitFields": validation.percentage_over_limit_fields,
                        },
                    },
                    status=400,
                )

            result = perform_calculation(payload)
            if "error" in result:
                return JsonResponse(
                    {"success": False, "message": result["error"]}, status=400
                )

            return JsonResponse({"success": True, "result": result})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=500)
